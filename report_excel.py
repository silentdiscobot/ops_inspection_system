# -*- coding: utf-8 -*-
import os
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from config import REPORT_DIR, CPU_THRESHOLD, MEM_THRESHOLD, DISK_THRESHOLD
from report_utils import build_report_filename

def over_limit_partitions(row):
    return [p for p in row.get("disk_partitions", []) if p.get("usage", 0) > DISK_THRESHOLD]

def partition_text(row):
    return "\n".join(f'{p.get("mount", "?")} ({p.get("usage", 0)}%)' for p in over_limit_partitions(row)) or "-"

def set_column_auto_width(ws):
    """自动调整列宽"""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max(max_length + 2, 10), 30)
        ws.column_dimensions[col_letter].width = adjusted_width

def generate_excel_report(project_name: str, inspector: str, date_str: str, rows: List[Dict[str, Any]], proxy_results: List[Dict[str, Any]] = None, check_cpu: bool = True, check_mem: bool = True, check_disk: bool = True, task_name: str = None) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "服务器巡检报告"

    # 定义样式
    title_font = Font(bold=True, size=16, color="FF1a365d")
    header_font = Font(bold=True, size=10, color="FFFFFFFF")
    normal_font = Font(size=10)
    bold_font = Font(bold=True, size=10)
    red_font = Font(bold=True, size=10, color="FFdc2626")
    green_font = Font(bold=True, size=10, color="FF059669")
    orange_font = Font(bold=True, size=10, color="FFea580c")
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style="thin", color="FFd1d5db"),
        right=Side(style="thin", color="FFd1d5db"),
        top=Side(style="thin", color="FFd1d5db"),
        bottom=Side(style="thin", color="FFd1d5db")
    )
    
    header_fill = PatternFill(start_color="FF1e40af", end_color="FF1e40af", fill_type="solid")
    proxy_header_fill = PatternFill(start_color="FF059669", end_color="FF059669", fill_type="solid")
    success_fill = PatternFill(start_color="FFd1fae5", end_color="FFd1fae5", fill_type="solid")
    warning_fill = PatternFill(start_color="FFFFf3cd", end_color="FFFFf3cd", fill_type="solid")
    danger_fill = PatternFill(start_color="FFfee2e2", end_color="FFfee2e2", fill_type="solid")

    # ========== 报告标题 ==========
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "服务器巡检报告"
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    
    ws.append([])
    
    # ========== 基本信息 ==========
    info_labels = [
        ("项目名称", project_name),
        ("巡检人", inspector),
        ("巡检时间", date_str),
        ("检查数量", f"{len(rows)} 台")
    ]
    
    for label, value in info_labels:
        ws.append([label, value])
        ws.cell(ws.max_row, 1).font = bold_font
        ws.cell(ws.max_row, 2).font = normal_font
    
    ws.append([])
    
    # ========== 服务器巡检记录表头（动态生成） ==========
    headers = ["序号", "服务器IP", "系统运行时间"]
    if check_cpu:
        headers.append("CPU使用率")
    if check_mem:
        headers.append("内存使用率")
    if check_disk:
        headers.append("磁盘使用率")
        headers.append("超限分区")
    headers.append("状态")
    ws.append(headers)
    
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(ws.max_row, c)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill

    # ========== 巡检数据行 ==========
    success_count = 0
    fail_count = 0
    warning_count = 0
    
    for idx, row in enumerate(rows, start=1):
        is_ok = row.get("ok", False)
        
        if not is_ok:
            status = "连接失败"
            status_font = red_font
            row_fill = danger_fill
            fail_count += 1
        else:
            issues = []
            if check_cpu and row.get("cpu", 0) > CPU_THRESHOLD:
                issues.append("CPU")
            if check_mem and row.get("mem", 0) > MEM_THRESHOLD:
                issues.append("内存")
            if check_disk and row.get("disk", 0) > DISK_THRESHOLD:
                issues.append("磁盘")
            
            if issues:
                status = f"告警({','.join(issues)})"
                status_font = orange_font
                row_fill = warning_fill
                warning_count += 1
            else:
                status = "正常"
                status_font = green_font
                row_fill = success_fill
                success_count += 1
        
        data = [
            idx,
            row.get("ip", ""),
            row.get("uptime", "未知"),
        ]
        if check_cpu:
            data.append(f'{row.get("cpu", 0)}%')
        if check_mem:
            data.append(f'{row.get("mem", 0)}%')
        if check_disk:
            data.append(f'{row.get("disk", 0)}%')
            data.append(partition_text(row))
        data.append(status)
        
        ws.append(data)
        for c in range(1, len(headers)+1):
            cell = ws.cell(ws.max_row, c)
            cell.alignment = center_alignment
            if check_disk and headers[c - 1] == "超限分区":
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = thin_border
            cell.fill = row_fill
            if c == len(headers):
                cell.font = status_font
    
    ws.append([])
    
    # ========== 网关代理检测结果 ==========
    if proxy_results:
        ws.append([])
        
        # 网关代理检测标题
        ws.merge_cells("A{}:F{}".format(ws.max_row + 1, ws.max_row + 1))
        proxy_title = ws.cell(ws.max_row + 1, 1)
        proxy_title.value = "网关代理检测结果"
        proxy_title.font = title_font
        proxy_title.alignment = center_alignment
        
        ws.append([])
        
        # 网关代理检测表头
        proxy_headers = ["序号", "服务器IP", "代理状态", "检测结果"]
        ws.append(proxy_headers)
        
        for c, h in enumerate(proxy_headers, start=1):
            cell = ws.cell(ws.max_row, c)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
            cell.fill = proxy_header_fill
        
        # 网关代理检测数据行
        proxy_success_count = 0
        proxy_fail_count = 0
        
        for idx, result in enumerate(proxy_results, start=1):
            if result.get("success"):
                status = "正常"
                status_font = green_font
                row_fill = success_fill
                proxy_success_count += 1
            elif result.get("error"):
                status = "连接失败"
                status_font = red_font
                row_fill = danger_fill
                proxy_fail_count += 1
            else:
                status = "异常"
                status_font = orange_font
                row_fill = warning_fill
                proxy_fail_count += 1
            
            output = result.get("output", "")
            if len(output) > 50:
                output = output[:50] + "..."
            
            data = [
                idx,
                result.get("ip", ""),
                status,
                output if output else (result.get("error", "无输出") if result.get("error") else "未包含成功关键词")
            ]
            
            ws.append(data)
            for c in range(1, len(proxy_headers)+1):
                cell = ws.cell(ws.max_row, c)
                cell.alignment = center_alignment
                cell.border = thin_border
                cell.fill = row_fill
                if c == 3:
                    cell.font = status_font
        
        ws.append([])
        
        # 网关代理检测汇总
        ws.append(["代理检测总数", f"{len(proxy_results)} 台"])
        ws.cell(ws.max_row, 1).font = bold_font
        ws.cell(ws.max_row, 2).font = normal_font
        
        ws.append(["代理正常", f"{proxy_success_count} 台"])
        ws.cell(ws.max_row, 1).font = bold_font
        ws.cell(ws.max_row, 2).font = green_font
        
        ws.append(["代理异常", f"{proxy_fail_count} 台"])
        ws.cell(ws.max_row, 1).font = bold_font
        ws.cell(ws.max_row, 2).font = red_font if proxy_fail_count > 0 else normal_font
    
    ws.append([])
    
    # ========== 统计汇总 ==========
    ws.merge_cells("A{}:F{}".format(ws.max_row + 1, ws.max_row + 1))
    summary_title = ws.cell(ws.max_row + 1, 1)
    summary_title.value = "巡检结果汇总"
    summary_title.font = bold_font
    summary_title.alignment = center_alignment
    
    summary_rows = [
        ("巡检总数", f"{len(rows)} 台"),
        ("正常运行", f"{success_count} 台"),
        ("告警警告", f"{warning_count} 台"),
        ("连接失败", f"{fail_count} 台"),
        ("巡检成功率", f"{round(success_count/len(rows)*100, 1) if rows else 0}%")
    ]
    
    for label, value in summary_rows:
        ws.append([label, value])
        ws.cell(ws.max_row, 1).font = normal_font
        ws.cell(ws.max_row, 2).font = bold_font
    
    ws.append([])
    
    # ========== 异常问题详情 ==========
    ws.merge_cells("A{}:F{}".format(ws.max_row + 1, ws.max_row + 1))
    issue_title = ws.cell(ws.max_row + 1, 1)
    issue_title.value = "异常问题详情"
    issue_title.font = bold_font
    issue_title.alignment = center_alignment
    
    abnormal_lines = []
    for row in rows:
        reasons = []
        if not row.get("ok"):
            reasons.append(f'连接失败: {row.get("error", "未知错误")}')
        else:
            if check_cpu and row.get("cpu", 0) > CPU_THRESHOLD:
                reasons.append(f'CPU使用率过高 {row["cpu"]}%')
            if check_mem and row.get("mem", 0) > MEM_THRESHOLD:
                reasons.append(f'内存使用率过高 {row["mem"]}%')
            if check_disk and row.get("disk", 0) > DISK_THRESHOLD:
                details = partition_text(row).replace("\n", "、")
                reasons.append(f'磁盘分区使用率过高: {details}')
        
        if reasons:
            abnormal_lines.append(f"{row.get('ip', '未知IP')}: {'; '.join(reasons)}")
    
    # 添加网关代理异常
    if proxy_results:
        for result in proxy_results:
            if not result.get("success"):
                ip = result.get("ip", "未知IP")
                if result.get("error"):
                    abnormal_lines.append(f"{ip}: 网关代理连接失败 - {result['error']}")
                else:
                    abnormal_lines.append(f"{ip}: 网关代理异常 - 未检测到成功关键词")
    
    if abnormal_lines:
        for line in abnormal_lines:
            ws.append([line])
            ws.cell(ws.max_row, 1).font = red_font
            ws.merge_cells(f"A{ws.max_row}:F{ws.max_row}")
    else:
        ws.append(["本次巡检未发现异常问题"])
        ws.cell(ws.max_row, 1).font = green_font
        ws.merge_cells(f"A{ws.max_row}:F{ws.max_row}")
    
    ws.append([])
    ws.append([])
    
    # ========== 签字区域 ==========
    ws.append(["巡检人：", "", "审核人：", "", "日期：", ""])
    ws.append(["", "", "", "", "", ""])
    
    for c in range(1, 7):
        cell_below = ws.cell(ws.max_row, c)
        cell_below.border = Border(bottom=Side(style="thin", color="FF4a5568"))
    
    # 自动调整列宽
    set_column_auto_width(ws)
    
    # 保存
    os.makedirs(REPORT_DIR, exist_ok=True)
    fname = build_report_filename(task_name or project_name, "xlsx")
    out_path = os.path.join(REPORT_DIR, fname)
    wb.save(out_path)
    return out_path
